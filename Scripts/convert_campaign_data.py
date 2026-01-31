#!/usr/bin/env python3
"""
Convert Campaign data.xlsx into CampaignData.json for the app bundle.
Run: python3 Scripts/convert_campaign_data.py
"""

import json
import openpyxl
from pathlib import Path

XLSX_PATH = Path(__file__).parent.parent / "Campaign data.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "ApPolForgeSym" / "CampaignData.json"


def read_sheet(wb, name):
    ws = wb[name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    return str(val) if val is not None else default


def build_state_data(wb):
    # Read all sheets
    type_action = {r["State"]: r for r in read_sheet(wb, "Type Action")}
    wk_last_20 = {r["State"]: r for r in read_sheet(wb, "Wk last 20")}
    roi = {r["State"]: r for r in read_sheet(wb, "ROI Investment")}
    hist = {r["State"]: r for r in read_sheet(wb, "Hist dat")}
    voter = {r["State"]: r for r in read_sheet(wb, "Voter Data")}
    staff_type = {r["State"]: r for r in read_sheet(wb, "Staff Type")}
    volunteers = {r["State"]: r for r in read_sheet(wb, "Volunteers")}
    budget = {r["State"]: r for r in read_sheet(wb, "budget Alllocations")}
    combo = {r["State"]: r for r in read_sheet(wb, "Combo1")}

    # Use Wk last 20 as canonical state list (has all 50 + abbreviation + EV + tier)
    states = []
    all_state_names = sorted(wk_last_20.keys())

    for state_name in all_state_names:
        wk = wk_last_20[state_name]
        ta = type_action.get(state_name, {})
        r = roi.get(state_name, {})
        h = hist.get(state_name, {})
        v = voter.get(state_name, {})
        st = staff_type.get(state_name, {})
        vol = volunteers.get(state_name, {})
        b = budget.get(state_name, {})
        c = combo.get(state_name, {})

        abbreviation = safe_str(wk.get("Abbreviation", ""))
        electoral_votes = safe_int(wk.get("Electoral_Votes", h.get("Electoral_Votes_2024", 0)))
        tier = safe_int(wk.get("Tier", ta.get("Competitiveness Tier", 4)))
        region = safe_str(ta.get("Region", c.get("Region", "Unknown")))

        # Historical data
        historical = {
            "winner2020": safe_str(h.get("Winner_2020")),
            "winner2016": safe_str(h.get("Winner_2016")),
            "winner2012": safe_str(h.get("Winner_2012")),
            "winner2008": safe_str(h.get("Winner_2008")),
            "margin2020": safe_float(h.get("Margin_2020_Pct")),
            "margin2016": safe_float(h.get("Margin_2016_Pct")),
            "margin2012": safe_float(h.get("Margin_2012_Pct")),
            "margin2008": safe_float(h.get("Margin_2008_Pct")),
            "trend": safe_str(h.get("Trend_2008_2020")),
            "turnout2020": safe_float(h.get("Turnout_2020_Pct")),
            "turnout2016": safe_float(h.get("Turnout_2016_Pct")),
        }

        # Action effectiveness (1-3 scores)
        action_effectiveness = {
            "townHall": safe_int(ta.get("Town Halls", 2)),
            "adCampaign": safe_int(ta.get("Ad Campaigns", 2)),
            "debate": safe_int(ta.get("Debate Prep", 1)),
            "rally": safe_int(ta.get("Rallies", 2)),
            "opposition": safe_int(ta.get("Opposition Research", 1)),
            "grassroots": safe_int(ta.get("Grassroots", 2)),
            "fundraiser": safe_int(ta.get("Fundraising", 2)),
        }

        # ROI data
        roi_data = {
            "swingPotentialScore": safe_int(r.get("Swing_Potential_Score", 15)),
            "roiRating": safe_str(r.get("ROI_Rating", "Low")),
            "spendEfficiencyRating": safe_str(r.get("Spend_Efficiency_Rating", "Low")),
            "costPerEV": safe_float(r.get("Cost_Per_EV_M", 0)),
            "totalSpend2020M": safe_float(r.get("Total_Spend_2020_M", 0)),
            "mediaMarketCostIndex": safe_float(b.get("Media_Market_Cost_Index", 1.0)),
        }

        # Staffing data
        staffing = {
            "totalStaff": safe_int(st.get("Total_Staff", 10)),
            "stateLeadership": safe_int(st.get("State_Leadership", 1)),
            "fieldOrganizers": safe_int(st.get("Field_Organizers", 3)),
            "communicationsStaff": safe_int(st.get("Communications_Staff", 1)),
            "regionalOffices": safe_int(v.get("Regional_Offices", c.get("Regional_Offices", 1))),
            "activeVolunteersPeak": safe_int(vol.get("Active_Volunteers_Peak", 1000)),
            "volunteerShiftsFinalMonth": safe_int(vol.get("Volunteer_Shifts_Final_Month", 3000)),
            "registeredVoters": safe_int(v.get("Estimated_Registered_Voters", vol.get("Registered_Voters", 0))),
        }

        # Budget data
        budget_data = {
            "totalBudgetM": safe_float(b.get("Total_State_Budget_Millions", 1.0)),
            "staffPayrollM": safe_float(b.get("Staff_Payroll_Millions", 0.1)),
            "tvAdvertisingM": safe_float(b.get("TV_Advertising_Millions", 0.1)),
            "digitalAdvertisingM": safe_float(b.get("Digital_Advertising_Millions", 0.05)),
            "gotvOperationsM": safe_float(b.get("GOTV_Operations_Millions", 0.05)),
            "earlyVoteInvestmentPct": safe_float(b.get("Early_Vote_Investment_Pct", 20)),
        }

        # Weekly pacing (20 weeks)
        weekly_pacing = []
        for w in range(20, 0, -1):
            weekly_pacing.append({
                "week": 21 - w,  # Convert: Wk20 -> week 1, Wk1 -> week 20
                "staff": safe_int(wk.get(f"Staff_Wk{w}")),
                "volunteers": safe_int(wk.get(f"Vol_Wk{w}")),
                "budgetK": safe_float(wk.get(f"Budget_Wk{w}_K")),
            })

        state_entry = {
            "name": state_name,
            "abbreviation": abbreviation,
            "electoralVotes": electoral_votes,
            "region": region,
            "competitivenessTier": tier,
            "historical": historical,
            "actionEffectiveness": action_effectiveness,
            "roi": roi_data,
            "staffing": staffing,
            "budget": budget_data,
            "weeklyPacing": weekly_pacing,
        }
        states.append(state_entry)

    # Compute metadata
    total_ev = sum(s["electoralVotes"] for s in states)
    total_budget = sum(s["budget"]["totalBudgetM"] for s in states)

    return {
        "metadata": {
            "totalElectoralVotes": total_ev,
            "totalBudgetAllStatesM": round(total_budget, 2),
            "stateCount": len(states),
        },
        "states": states,
    }


def main():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    data = build_state_data(wb)

    with open(str(OUTPUT_PATH), "w") as f:
        json.dump(data, f, indent=2)

    print(f"Generated {OUTPUT_PATH}")
    print(f"  States: {data['metadata']['stateCount']}")
    print(f"  Total EV: {data['metadata']['totalElectoralVotes']}")
    print(f"  Total Budget: ${data['metadata']['totalBudgetAllStatesM']}M")


if __name__ == "__main__":
    main()
